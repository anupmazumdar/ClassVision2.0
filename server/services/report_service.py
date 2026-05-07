"""PDF + Excel report generation."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import os

def generate_excel(session_data, output_path):
    wb  = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Attendance"

    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="1a1a2e")
    headers  = ["Enrollment", "Name", "Status", "Time", "Confidence %"]
    for c, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=c, value=h)
        cell.font = hdr_font; cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center")

    for i, rec in enumerate(session_data.get("attendance",[]), 2):
        ws1.cell(row=i, column=1, value=rec.get("enrollment"))
        ws1.cell(row=i, column=2, value=rec.get("name"))
        ws1.cell(row=i, column=3, value="Present")
        ws1.cell(row=i, column=4, value=str(rec.get("timestamp","")))
        ws1.cell(row=i, column=5, value=rec.get("confidence"))

    for c in range(1,6): ws1.column_dimensions[get_column_letter(c)].width = 20

    ws2 = wb.create_sheet("Emotions")
    ws2.append(["Emotion","Percentage"])
    for e,p in session_data.get("emotions",{}).items(): ws2.append([e.capitalize(), f"{p}%"])

    ws3 = wb.create_sheet("Attention")
    ws3.append(["Class Attention %", session_data.get("attention_pct","N/A")])

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    return output_path
