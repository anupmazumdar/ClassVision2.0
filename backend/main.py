from fastapi import FastAPI, Depends, HTTPException, status
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, EmailStr
from sqlalchemy.orm import Session
from typing import List, Optional
from datetime import datetime
import uvicorn, os, json, io

from database import get_db, init_db, User, Student, ClassSession, AttendanceRecord
from auth import (
    hash_password, verify_password, create_token, get_current_user, require_admin
)
try:
    from face_service import decode_image, extract_encodings, recognize_faces
except ImportError:
    from face_service_mock import decode_image, extract_encodings, recognize_faces

app = FastAPI(title="ClassVision API", version="2.0.0")
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.on_event("startup")
def startup():
    init_db()
    _seed_admin()


def _seed_admin():
    """Create a default admin on first run if no users exist."""
    from database import SessionLocal
    db = SessionLocal()
    try:
        if db.query(User).count() == 0:
            db.add(User(
                name="Admin",
                email=os.getenv("ADMIN_EMAIL", "admin@classvision.local"),
                password_hash=hash_password(os.getenv("ADMIN_PASSWORD", "admin123")),
                role="admin",
            ))
            db.commit()
    finally:
        db.close()


# ── Health ────────────────────────────────────────────────────────────────────

@app.get("/")
def root():
    return {"status": "ClassVision running", "version": "2.0.0"}


# ── Auth ──────────────────────────────────────────────────────────────────────

class LoginRequest(BaseModel):
    email: str
    password: str


class RegisterRequest(BaseModel):
    name: str
    email: str
    password: str
    role: str = "teacher"


@app.post("/auth/login")
def login(body: LoginRequest, db: Session = Depends(get_db)):
    user = db.query(User).filter(User.email == body.email).first()
    if not user or not verify_password(body.password, user.password_hash):
        raise HTTPException(status_code=401, detail="Invalid email or password")
    token = create_token(user.id, user.email, user.role, user.name)
    return {"access_token": token, "token_type": "bearer", "role": user.role, "name": user.name}


@app.post("/auth/register", status_code=201)
def register(body: RegisterRequest, db: Session = Depends(get_db), _=Depends(require_admin)):
    if db.query(User).filter(User.email == body.email).first():
        raise HTTPException(status_code=400, detail="Email already registered")
    user = User(
        name=body.name,
        email=body.email,
        password_hash=hash_password(body.password),
        role=body.role,
    )
    db.add(user)
    db.commit()
    db.refresh(user)
    return {"id": user.id, "name": user.name, "email": user.email, "role": user.role}


@app.get("/auth/me")
def me(current: dict = Depends(get_current_user)):
    return current


# ── Students ──────────────────────────────────────────────────────────────────

class StudentCreate(BaseModel):
    enrollment: str
    name: str
    department: str = ""


class FaceRegisterRequest(BaseModel):
    images: List[str]  # list of base64-encoded images


@app.get("/students")
def list_students(db: Session = Depends(get_db), _=Depends(get_current_user)):
    students = db.query(Student).order_by(Student.name).all()
    return [
        {
            "id": s.id,
            "enrollment": s.enrollment,
            "name": s.name,
            "department": s.department,
            "has_face": bool(json.loads(s.face_encodings or "[]")),
            "created_at": s.created_at.isoformat(),
        }
        for s in students
    ]


@app.post("/students", status_code=201)
def create_student(body: StudentCreate, db: Session = Depends(get_db), _=Depends(get_current_user)):
    if db.query(Student).filter(Student.enrollment == body.enrollment).first():
        raise HTTPException(status_code=400, detail="Enrollment number already exists")
    student = Student(enrollment=body.enrollment, name=body.name, department=body.department)
    db.add(student)
    db.commit()
    db.refresh(student)
    return {"id": student.id, "enrollment": student.enrollment, "name": student.name}


@app.post("/students/{student_id}/register-face")
def register_face(
    student_id: int,
    body: FaceRegisterRequest,
    db: Session = Depends(get_db),
    _=Depends(get_current_user),
):
    student = db.query(Student).filter(Student.id == student_id).first()
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")

    all_encodings = []
    for b64 in body.images:
        try:
            img_array = decode_image(b64)
            encs = extract_encodings(img_array)
            all_encodings.extend(encs)
        except Exception as e:
            continue

    if not all_encodings:
        raise HTTPException(status_code=422, detail="No face detected in the provided image(s). Please use a clear, well-lit photo.")

    student.face_encodings = json.dumps(all_encodings)
    db.commit()
    return {"message": f"Registered {len(all_encodings)} face encoding(s) for {student.name}"}


@app.delete("/students/{student_id}", status_code=204)
def delete_student(student_id: int, db: Session = Depends(get_db), _=Depends(require_admin)):
    student = db.query(Student).filter(Student.id == student_id).first()
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")
    db.query(AttendanceRecord).filter(AttendanceRecord.student_id == student_id).delete()
    db.delete(student)
    db.commit()


# ── Sessions ──────────────────────────────────────────────────────────────────

class SessionCreate(BaseModel):
    subject: str
    room: str = ""


@app.get("/sessions")
def list_sessions(db: Session = Depends(get_db), current: dict = Depends(get_current_user)):
    sessions = (
        db.query(ClassSession)
        .order_by(ClassSession.started_at.desc())
        .limit(50)
        .all()
    )
    result = []
    for s in sessions:
        count = db.query(AttendanceRecord).filter(AttendanceRecord.session_id == s.id).count()
        result.append({
            "id": s.id,
            "subject": s.subject,
            "room": s.room,
            "is_active": s.is_active,
            "started_at": s.started_at.isoformat(),
            "ended_at": s.ended_at.isoformat() if s.ended_at else None,
            "present_count": count,
        })
    return result


@app.post("/sessions", status_code=201)
def start_session(body: SessionCreate, db: Session = Depends(get_db), current: dict = Depends(get_current_user)):
    # Close any previously active session for this teacher
    db.query(ClassSession).filter(
        ClassSession.teacher_id == int(current["sub"]),
        ClassSession.is_active == True,
    ).update({"is_active": False, "ended_at": datetime.utcnow()})

    session = ClassSession(
        subject=body.subject,
        room=body.room,
        teacher_id=int(current["sub"]),
    )
    db.add(session)
    db.commit()
    db.refresh(session)
    return {"id": session.id, "subject": session.subject, "started_at": session.started_at.isoformat()}


@app.delete("/sessions/{session_id}", status_code=204)
def delete_session(session_id: int, db: Session = Depends(get_db), _=Depends(require_admin)):
    session = db.query(ClassSession).filter(ClassSession.id == session_id).first()
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")
    db.query(AttendanceRecord).filter(AttendanceRecord.session_id == session_id).delete()
    db.delete(session)
    db.commit()


@app.put("/sessions/{session_id}/stop")
def stop_session(session_id: int, db: Session = Depends(get_db), _=Depends(get_current_user)):
    session = db.query(ClassSession).filter(ClassSession.id == session_id).first()
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")
    session.is_active = False
    session.ended_at = datetime.utcnow()
    db.commit()
    count = db.query(AttendanceRecord).filter(AttendanceRecord.session_id == session_id).count()
    return {"message": "Session ended", "present_count": count}


@app.get("/sessions/{session_id}")
def get_session(session_id: int, db: Session = Depends(get_db), _=Depends(get_current_user)):
    session = db.query(ClassSession).filter(ClassSession.id == session_id).first()
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")

    records = (
        db.query(AttendanceRecord, Student)
        .join(Student, AttendanceRecord.student_id == Student.id)
        .filter(AttendanceRecord.session_id == session_id)
        .all()
    )
    attendance = [
        {
            "student_id": s.id,
            "enrollment": s.enrollment,
            "name": s.name,
            "department": s.department,
            "confidence": r.confidence,
            "marked_at": r.marked_at.isoformat(),
        }
        for r, s in records
    ]

    return {
        "id": session.id,
        "subject": session.subject,
        "room": session.room,
        "is_active": session.is_active,
        "started_at": session.started_at.isoformat(),
        "ended_at": session.ended_at.isoformat() if session.ended_at else None,
        "attendance": attendance,
    }


# ── Attendance ────────────────────────────────────────────────────────────────

class RecognizeRequest(BaseModel):
    image: str  # base64 frame


class MarkRequest(BaseModel):
    student_id: int
    confidence: float = 0.0


@app.post("/attendance/recognize")
def recognize(body: RecognizeRequest, db: Session = Depends(get_db), _=Depends(get_current_user)):
    """Send a camera frame, get back list of recognized students."""
    try:
        img = decode_image(body.image)
    except Exception:
        raise HTTPException(status_code=422, detail="Invalid image data")

    students = db.query(Student).filter(Student.face_encodings != "[]").all()
    if not students:
        return {"recognized": [], "message": "No students with registered faces"}

    recognized = recognize_faces(img, students)
    return {"recognized": recognized}


@app.post("/attendance/{session_id}/mark", status_code=201)
def mark_attendance(
    session_id: int,
    body: MarkRequest,
    db: Session = Depends(get_db),
    _=Depends(get_current_user),
):
    session = db.query(ClassSession).filter(ClassSession.id == session_id).first()
    if not session or not session.is_active:
        raise HTTPException(status_code=400, detail="Session not found or not active")

    # Prevent duplicate marks in same session
    existing = db.query(AttendanceRecord).filter(
        AttendanceRecord.session_id == session_id,
        AttendanceRecord.student_id == body.student_id,
    ).first()
    if existing:
        return {"message": "Already marked", "already_present": True}

    record = AttendanceRecord(
        session_id=session_id,
        student_id=body.student_id,
        confidence=body.confidence,
    )
    db.add(record)
    db.commit()
    return {"message": "Marked present", "already_present": False}


# ── Reports ───────────────────────────────────────────────────────────────────

# ── Users (admin) ─────────────────────────────────────────────────────────────

@app.get("/users")
def list_users(db: Session = Depends(get_db), _=Depends(require_admin)):
    users = db.query(User).order_by(User.name).all()
    return [
        {"id": u.id, "name": u.name, "email": u.email, "role": u.role, "created_at": u.created_at.isoformat()}
        for u in users
    ]


@app.delete("/users/{user_id}", status_code=204)
def delete_user(user_id: int, db: Session = Depends(get_db), current: dict = Depends(require_admin)):
    if str(user_id) == current["sub"]:
        raise HTTPException(status_code=400, detail="Cannot delete your own account")
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    db.delete(user)
    db.commit()


# ── Attendance: manual unmark ─────────────────────────────────────────────────

@app.delete("/attendance/{session_id}/unmark/{student_id}", status_code=204)
def unmark_attendance(session_id: int, student_id: int, db: Session = Depends(get_db), _=Depends(get_current_user)):
    record = db.query(AttendanceRecord).filter(
        AttendanceRecord.session_id == session_id,
        AttendanceRecord.student_id == student_id,
    ).first()
    if record:
        db.delete(record)
        db.commit()


# ── Reports: student-wise summary ────────────────────────────────────────────

@app.get("/reports/student-summary")
def student_summary(db: Session = Depends(get_db), _=Depends(get_current_user)):
    total = db.query(ClassSession).filter(ClassSession.is_active == False).count()
    students = db.query(Student).order_by(Student.name).all()
    rows = []
    for s in students:
        present = db.query(AttendanceRecord).filter(AttendanceRecord.student_id == s.id).count()
        rows.append({
            "id": s.id,
            "name": s.name,
            "enrollment": s.enrollment,
            "department": s.department,
            "present": present,
            "total": total,
            "pct": round(present / total * 100, 1) if total > 0 else 0,
        })
    return {"total_sessions": total, "students": sorted(rows, key=lambda x: x["pct"], reverse=True)}


# ── Reports: PDF ──────────────────────────────────────────────────────────────

@app.get("/reports/{session_id}/pdf")
def export_pdf(session_id: int, db: Session = Depends(get_db), _=Depends(get_current_user)):
    try:
        from fpdf import FPDF
    except ImportError:
        raise HTTPException(status_code=500, detail="fpdf2 not installed")

    session = db.query(ClassSession).filter(ClassSession.id == session_id).first()
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")

    records = (
        db.query(AttendanceRecord, Student)
        .join(Student, AttendanceRecord.student_id == Student.id)
        .filter(AttendanceRecord.session_id == session_id)
        .order_by(Student.name)
        .all()
    )
    total_students = db.query(Student).count()

    pdf = FPDF()
    pdf.add_page()

    pdf.set_fill_color(30, 58, 95)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font("Helvetica", "B", 17)
    pdf.cell(0, 13, "ClassVision — Attendance Report", fill=True, align="C", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(5)

    pdf.set_text_color(40, 40, 40)
    info = [
        ("Subject", session.subject),
        ("Room", session.room or "—"),
        ("Date", session.started_at.strftime("%d %B %Y")),
        ("Time", session.started_at.strftime("%H:%M") + (" — " + session.ended_at.strftime("%H:%M") if session.ended_at else "")),
        ("Present", f"{len(records)} / {total_students}  ({round(len(records)/max(total_students,1)*100,1)}%)"),
    ]
    for label, val in info:
        pdf.set_font("Helvetica", "B", 10)
        pdf.cell(38, 8, label + ":", new_x="RIGHT", new_y="TOP")
        pdf.set_font("Helvetica", "", 10)
        pdf.cell(0, 8, val, new_x="LMARGIN", new_y="NEXT")
    pdf.ln(4)

    # Table header
    pdf.set_fill_color(30, 58, 95)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font("Helvetica", "B", 9)
    col_w = [10, 42, 72, 40, 26]
    for txt, w in zip(["#", "Enrollment", "Name", "Department", "Conf%"], col_w):
        pdf.cell(w, 8, txt, fill=True, border=1, align="C")
    pdf.ln()

    pdf.set_text_color(30, 30, 30)
    pdf.set_font("Helvetica", "", 9)
    for i, (r, s) in enumerate(records, 1):
        pdf.set_fill_color(245, 247, 252) if i % 2 == 0 else pdf.set_fill_color(255, 255, 255)
        for txt, w in zip([str(i), s.enrollment, s.name, s.department or "—", f"{round(r.confidence,1)}%"], col_w):
            pdf.cell(w, 7, txt, fill=True, border=1)
        pdf.ln()

    buf = io.BytesIO(pdf.output())
    fname = f"attendance_{session.subject}_{session.started_at.strftime('%Y%m%d_%H%M')}.pdf"
    return StreamingResponse(buf, media_type="application/pdf",
                             headers={"Content-Disposition": f'attachment; filename="{fname}"'})


# ── Reports: email ────────────────────────────────────────────────────────────

class EmailRequest(BaseModel):
    to: str


@app.post("/reports/{session_id}/email")
def email_report(session_id: int, body: EmailRequest, db: Session = Depends(get_db), _=Depends(get_current_user)):
    import smtplib
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText
    from email.mime.base import MIMEBase
    from email import encoders as email_encoders

    smtp_host = os.getenv("SMTP_HOST", "")
    smtp_port = int(os.getenv("SMTP_PORT", "587"))
    smtp_user = os.getenv("SMTP_USER", "")
    smtp_pass = os.getenv("SMTP_PASS", "")
    smtp_from = os.getenv("SMTP_FROM", smtp_user)

    if not smtp_host or not smtp_user or not smtp_pass:
        raise HTTPException(status_code=503, detail="Email not configured. Add SMTP_HOST, SMTP_USER, SMTP_PASS to .env")

    session = db.query(ClassSession).filter(ClassSession.id == session_id).first()
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")

    records = (
        db.query(AttendanceRecord, Student)
        .join(Student, AttendanceRecord.student_id == Student.id)
        .filter(AttendanceRecord.session_id == session_id)
        .order_by(Student.name)
        .all()
    )

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"
    hfill = PatternFill("solid", fgColor="1e3a5f")
    hfont = Font(bold=True, color="FFFFFF")
    for col, h in enumerate(["#", "Enrollment", "Name", "Department", "Confidence %", "Marked At"], 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = hfont
        c.fill = hfill
        c.alignment = Alignment(horizontal="center")
    for i, (r, s) in enumerate(records, 1):
        ws.append([i, s.enrollment, s.name, s.department, round(r.confidence, 1), r.marked_at.strftime("%Y-%m-%d %H:%M:%S")])

    excel_buf = io.BytesIO()
    wb.save(excel_buf)
    excel_buf.seek(0)

    msg = MIMEMultipart()
    msg["From"] = smtp_from
    msg["To"] = body.to
    msg["Subject"] = f"Attendance — {session.subject} ({session.started_at.strftime('%d %b %Y')})"
    msg.attach(MIMEText(
        f"Subject: {session.subject}\nDate: {session.started_at.strftime('%d %B %Y')}\n"
        f"Present: {len(records)}\n\nAttendance report attached.\n\n— ClassVision",
        "plain"
    ))
    part = MIMEBase("application", "octet-stream")
    part.set_payload(excel_buf.read())
    email_encoders.encode_base64(part)
    fname = f"attendance_{session.subject}_{session.started_at.strftime('%Y%m%d_%H%M')}.xlsx"
    part.add_header("Content-Disposition", f'attachment; filename="{fname}"')
    msg.attach(part)

    try:
        with smtplib.SMTP(smtp_host, smtp_port, timeout=10) as srv:
            srv.ehlo()
            srv.starttls()
            srv.login(smtp_user, smtp_pass)
            srv.sendmail(smtp_from, body.to, msg.as_string())
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"Failed to send email: {e}")

    return {"message": f"Report sent to {body.to}"}


@app.get("/reports/{session_id}/excel")
def export_excel(session_id: int, db: Session = Depends(get_db), _=Depends(get_current_user)):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    session = db.query(ClassSession).filter(ClassSession.id == session_id).first()
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")

    records = (
        db.query(AttendanceRecord, Student)
        .join(Student, AttendanceRecord.student_id == Student.id)
        .filter(AttendanceRecord.session_id == session_id)
        .order_by(Student.name)
        .all()
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"

    header_fill = PatternFill("solid", fgColor="1e3a5f")
    header_font = Font(bold=True, color="FFFFFF")

    headers = ["#", "Enrollment", "Name", "Department", "Confidence %", "Marked At"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for i, (r, s) in enumerate(records, 1):
        ws.append([
            i,
            s.enrollment,
            s.name,
            s.department,
            round(r.confidence, 1),
            r.marked_at.strftime("%Y-%m-%d %H:%M:%S"),
        ])

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    total_students = db.query(Student).count()
    ws2.append(["Subject", session.subject])
    ws2.append(["Room", session.room])
    ws2.append(["Date", session.started_at.strftime("%Y-%m-%d")])
    ws2.append(["Start Time", session.started_at.strftime("%H:%M:%S")])
    ws2.append(["End Time", session.ended_at.strftime("%H:%M:%S") if session.ended_at else "—"])
    ws2.append(["Total Students", total_students])
    ws2.append(["Present", len(records)])
    ws2.append(["Absent", max(0, total_students - len(records))])
    ws2.append(["Attendance %", f"{round(len(records) / max(total_students, 1) * 100, 1)}%"])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"attendance_{session.subject}_{session.started_at.strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


if __name__ == "__main__":
    uvicorn.run(
        "main:app",
        host=os.getenv("SERVER_HOST", "0.0.0.0"),
        port=int(os.getenv("SERVER_PORT", "8000")),
        reload=True,
    )
