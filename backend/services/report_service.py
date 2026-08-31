import io
import os
import smtplib
from email import encoders as email_encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

from fastapi import HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from repositories import attendance_repo, session_repo, student_repo


def get_student_summary(db: Session) -> dict:
    total = session_repo.count_closed_sessions(db)
    students = student_repo.list_students(db)
    rows = []
    for s in students:
        present = attendance_repo.count_by_student(db, s.id)
        rows.append(
            {
                "id": s.id,
                "name": s.name,
                "enrollment": s.enrollment,
                "department": s.department,
                "present": present,
                "total": total,
                "pct": round(present / total * 100, 1) if total > 0 else 0,
            }
        )

    return {"total_sessions": total, "students": sorted(rows, key=lambda x: x["pct"], reverse=True)}


def _get_session_with_records(db: Session, session_id: int):
    session = session_repo.get_session_by_id(db, session_id)
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")

    records = attendance_repo.list_session_records_with_students(db, session_id)
    return session, records


import re

def _sanitize_header_value(val: str) -> str:
    """Sanitizes text strings for safe usage in HTTP headers to prevent header injection / response splitting."""
    return re.sub(r"[^A-Za-z0-9_.-]", "_", str(val).strip())[:60]


def _sanitize_excel_cell(val):
    """Prevents CSV / Excel Formula Injection (CWE-1236) by quoting leading formula trigger characters."""
    if isinstance(val, str) and val.startswith(("=", "+", "-", "@")):
        return f"'{val}"
    return val


def export_pdf(db: Session, session_id: int):
    try:
        from fpdf import FPDF
    except ImportError:
        raise HTTPException(status_code=500, detail="fpdf2 not installed")

    session, records = _get_session_with_records(db, session_id)
    total_students = student_repo.count_students(db)

    pdf = FPDF()
    pdf.add_page()

    pdf.set_fill_color(30, 58, 95)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font("Helvetica", "B", 17)
    pdf.cell(0, 13, "ClassVision - Attendance Report", fill=True, align="C", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(5)

    pdf.set_text_color(40, 40, 40)
    info = [
        ("Subject", session.subject),
        ("Room", session.room or "-"),
        ("Date", session.started_at.strftime("%d %B %Y")),
        (
            "Time",
            session.started_at.strftime("%H:%M")
            + (" - " + session.ended_at.strftime("%H:%M") if session.ended_at else ""),
        ),
        (
            "Present",
            f"{len(records)} / {total_students}  ({round(len(records) / max(total_students, 1) * 100, 1)}%)",
        ),
    ]
    for label, val in info:
        pdf.set_font("Helvetica", "B", 10)
        pdf.cell(38, 8, label + ":", new_x="RIGHT", new_y="TOP")
        pdf.set_font("Helvetica", "", 10)
        pdf.cell(0, 8, str(val), new_x="LMARGIN", new_y="NEXT")
    pdf.ln(4)

    pdf.set_fill_color(30, 58, 95)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font("Helvetica", "B", 9)
    col_w = [10, 42, 72, 40, 26]
    for txt, width in zip(["#", "Enrollment", "Name", "Department", "Conf%"], col_w):
        pdf.cell(width, 8, txt, fill=True, border=1, align="C")
    pdf.ln()

    pdf.set_text_color(30, 30, 30)
    pdf.set_font("Helvetica", "", 9)
    for i, (record, student) in enumerate(records, 1):
        if i % 2 == 0:
            pdf.set_fill_color(245, 247, 252)
        else:
            pdf.set_fill_color(255, 255, 255)

        for txt, width in zip(
            [
                str(i),
                student.enrollment,
                student.name,
                student.department or "-",
                f"{round(record.confidence, 1)}%",
            ],
            col_w,
        ):
            pdf.cell(width, 7, txt, fill=True, border=1)
        pdf.ln()

    buf = io.BytesIO(pdf.output())
    safe_subject = _sanitize_header_value(session.subject)
    filename = f"attendance_{safe_subject}_{session.started_at.strftime('%Y%m%d_%H%M')}.pdf"
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def export_excel(db: Session, session_id: int):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    session, records = _get_session_with_records(db, session_id)

    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"

    header_fill = PatternFill("solid", fgColor="1e3a5f")
    header_font = Font(bold=True, color="FFFFFF")

    headers = ["#", "Enrollment", "Name", "Department", "Confidence %", "Marked At"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for i, (record, student) in enumerate(records, 1):
        ws.append(
            [
                i,
                _sanitize_excel_cell(student.enrollment),
                _sanitize_excel_cell(student.name),
                _sanitize_excel_cell(student.department or ""),
                round(record.confidence, 1),
                record.marked_at.strftime("%Y-%m-%d %H:%M:%S"),
            ]
        )

    ws2 = wb.create_sheet("Summary")
    total_students = student_repo.count_students(db)
    ws2.append(["Subject", _sanitize_excel_cell(session.subject)])
    ws2.append(["Room", _sanitize_excel_cell(session.room or "")])
    ws2.append(["Date", session.started_at.strftime("%Y-%m-%d")])
    ws2.append(["Start Time", session.started_at.strftime("%H:%M:%S")])
    ws2.append(["End Time", session.ended_at.strftime("%H:%M:%S") if session.ended_at else "-"])
    ws2.append(["Total Students", total_students])
    ws2.append(["Present", len(records)])
    ws2.append(["Absent", max(0, total_students - len(records))])
    ws2.append(["Attendance %", f"{round(len(records) / max(total_students, 1) * 100, 1)}%"])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    safe_subject = _sanitize_header_value(session.subject)
    filename = f"attendance_{safe_subject}_{session.started_at.strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def email_report(db: Session, session_id: int, to_email: str) -> dict:
    smtp_host = os.getenv("SMTP_HOST", "")
    smtp_port = int(os.getenv("SMTP_PORT", "587"))
    smtp_user = os.getenv("SMTP_USER", "")
    smtp_pass = os.getenv("SMTP_PASS", "")
    smtp_from = os.getenv("SMTP_FROM", smtp_user)

    if not smtp_host or not smtp_user or not smtp_pass:
        raise HTTPException(
            status_code=503,
            detail="Email not configured. Add SMTP_HOST, SMTP_USER, SMTP_PASS to .env",
        )

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    session, records = _get_session_with_records(db, session_id)

    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"

    hfill = PatternFill("solid", fgColor="1e3a5f")
    hfont = Font(bold=True, color="FFFFFF")
    for col, header in enumerate(["#", "Enrollment", "Name", "Department", "Confidence %", "Marked At"], 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = hfont
        cell.fill = hfill
        cell.alignment = Alignment(horizontal="center")

    for i, (record, student) in enumerate(records, 1):
        ws.append(
            [
                i,
                student.enrollment,
                student.name,
                student.department,
                round(record.confidence, 1),
                record.marked_at.strftime("%Y-%m-%d %H:%M:%S"),
            ]
        )

    excel_buf = io.BytesIO()
    wb.save(excel_buf)
    excel_buf.seek(0)

    message = MIMEMultipart()
    message["From"] = smtp_from
    message["To"] = to_email
    message["Subject"] = f"Attendance - {session.subject} ({session.started_at.strftime('%d %b %Y')})"
    message.attach(
        MIMEText(
            f"Subject: {session.subject}\nDate: {session.started_at.strftime('%d %B %Y')}\n"
            f"Present: {len(records)}\n\nAttendance report attached.\n\n- ClassVision",
            "plain",
        )
    )

    part = MIMEBase("application", "octet-stream")
    part.set_payload(excel_buf.read())
    email_encoders.encode_base64(part)
    filename = f"attendance_{session.subject}_{session.started_at.strftime('%Y%m%d_%H%M')}.xlsx"
    part.add_header("Content-Disposition", f'attachment; filename="{filename}"')
    message.attach(part)

    try:
        with smtplib.SMTP(smtp_host, smtp_port, timeout=10) as server:
            server.ehlo()
            server.starttls()
            server.login(smtp_user, smtp_pass)
            server.sendmail(smtp_from, to_email, message.as_string())
    except Exception as exc:
        raise HTTPException(status_code=502, detail=f"Failed to send email: {exc}")

    return {"message": f"Report sent to {to_email}"}
